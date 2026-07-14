"""Unit tests for report generation (Excel, HTML, JSON)."""

import json
from pathlib import Path

from openpyxl import load_workbook

from webval.reports import write_excel_report, write_html_report, write_json_results
from webval.reports.json_out import read_json_results


class TestExcelReport:
    def test_workbook_structure(self, sample_run, tmp_path: Path):
        out = write_excel_report(sample_run, tmp_path / "matrix.xlsx")
        wb = load_workbook(out)
        assert wb.sheetnames == [
            "Executive Summary", "Defect Log", "Traceability Matrix", "Evidence Index",
        ]

    def test_matrix_rows_cover_all_requirements(self, sample_run, tmp_path: Path):
        wb = load_workbook(write_excel_report(sample_run, tmp_path / "matrix.xlsx"))
        ws = wb["Traceability Matrix"]
        ids = [ws.cell(row=r, column=1).value for r in range(2, ws.max_row + 1)]
        assert ids == ["REQ-001", "REQ-002", "REQ-003", "REQ-004"]
        statuses = [ws.cell(row=r, column=6).value for r in range(2, ws.max_row + 1)]
        assert statuses == ["Pass", "Fail", "Warning", "Not Tested"]

    def test_summary_counts(self, sample_run, tmp_path: Path):
        wb = load_workbook(write_excel_report(sample_run, tmp_path / "matrix.xlsx"))
        ws = wb["Executive Summary"]
        values = {ws.cell(row=r, column=1).value: ws.cell(row=r, column=2).value
                  for r in range(1, ws.max_row + 1)}
        assert values["Total Requirements"] == "4"
        assert values["Passed"] == "1"
        assert values["Failed"] == "1"

    def test_defect_log_format(self, sample_run, tmp_path: Path):
        wb = load_workbook(write_excel_report(sample_run, tmp_path / "matrix.xlsx"))
        ws = wb["Defect Log"]
        headers = [ws.cell(row=1, column=c).value for c in range(1, 21)]
        assert headers == [
            "Defect No", "QA Name", "QA Date", "Phase", "Environment", "Round of Testing",
            "Defect URL", "Device", "Section / Page Name", "Steps / Defect Description",
            "Defect Type", "Defect Category", "Severity", "QA Status", "DEV Status",
            "Reference/Verified File", "Defect Screenshot", "Developer Comments",
            "QC Comments", "DPM Comments",
        ]
        # 2 defects in the fixture (1 Fail + 1 Warning)
        assert ws.cell(row=2, column=1).value == "DEF-001"
        assert ws.cell(row=2, column=13).value == "Major"
        assert ws.cell(row=3, column=13).value == "Minor"
        assert ws.cell(row=2, column=14).value == "Open"
        assert ws.cell(row=2, column=17).value == "evidence/screenshots/0001-fail.png"

    def test_evidence_index(self, sample_run, tmp_path: Path):
        wb = load_workbook(write_excel_report(sample_run, tmp_path / "matrix.xlsx"))
        ws = wb["Evidence Index"]
        assert ws.cell(row=2, column=1).value == "REQ-002"
        assert ws.cell(row=2, column=3).value == "evidence/screenshots/0001-fail.png"


class TestHtmlReport:
    def test_report_renders_all_sections(self, sample_run, tmp_path: Path):
        out = write_html_report(
            sample_run, tmp_path / "report.html",
            title="Validation Report", organization="QA", system_under_test="example.test",
        )
        html = out.read_text()
        for fragment in (
            "Executive Summary", "Traceability Matrix", "Defect Summary",
            "REQ-001", "REQ-004", "Not Tested", "evidence/screenshots/0001-fail.png",
        ):
            assert fragment in html

    def test_requirement_text_escaped(self, sample_run, tmp_path: Path):
        sample_run.requirements.requirements[0].requirement = 'XSS <script>alert(1)</script>'
        out = write_html_report(
            sample_run, tmp_path / "report.html",
            title="t", organization="o", system_under_test="s",
        )
        html = out.read_text()
        assert "<script>alert(1)</script>" not in html
        assert "&lt;script&gt;" in html


class TestJsonResults:
    def test_round_trip(self, sample_run, tmp_path: Path):
        out = write_json_results(sample_run, tmp_path / "results.json")
        loaded = read_json_results(out)
        assert loaded.manifest.run_id == sample_run.manifest.run_id
        assert len(loaded.results) == 3
        assert loaded.summary.total == 4

    def test_summary_and_defects_embedded(self, sample_run, tmp_path: Path):
        out = write_json_results(sample_run, tmp_path / "results.json")
        data = json.loads(out.read_text())
        assert data["summary"]["passed"] == 1
        assert len(data["defects"]) == 2
