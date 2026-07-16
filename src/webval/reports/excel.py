"""Excel traceability matrix (openpyxl).

Workbook layout:
  1. Executive Summary — run manifest + status counts + pass rate
  2. Traceability Matrix — one row per requirement (the auditable core)
  3. Defect Summary — Fail/Warning/Error rows condensed for triage
  4. Evidence Index — every ledgered artifact with hash and timestamp
"""

from __future__ import annotations

from pathlib import Path
from urllib.parse import urlparse

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet

from webval.config.settings import ReportConfig
from webval.models import EvidenceKind, Status, ValidationRun
from webval.utils import get_logger

log = get_logger("reports.excel")

_STATUS_FILL = {
    Status.PASS: PatternFill("solid", start_color="C6EFCE"),
    Status.FAIL: PatternFill("solid", start_color="FFC7CE"),
    Status.WARNING: PatternFill("solid", start_color="FFEB9C"),
    Status.NOT_TESTED: PatternFill("solid", start_color="D9D9D9"),
    Status.ERROR: PatternFill("solid", start_color="F4B084"),
}
_STATUS_FONT = {
    Status.PASS: Font(color="006100"),
    Status.FAIL: Font(color="9C0006", bold=True),
    Status.WARNING: Font(color="9C6500"),
    Status.NOT_TESTED: Font(color="595959"),
    Status.ERROR: Font(color="833C00", bold=True),
}
_HEADER_FILL = PatternFill("solid", start_color="1F4E79")
_HEADER_FONT = Font(color="FFFFFF", bold=True)
_THIN = Border(*(Side(style="thin", color="BFBFBF"),) * 4)


def write_excel_report(
    run: ValidationRun, output_path: Path, report_cfg: ReportConfig | None = None
) -> Path:
    wb = Workbook()
    summary_ws = wb.active
    assert isinstance(summary_ws, Worksheet)
    _summary_sheet(summary_ws, run)
    _defect_log_sheet(wb.create_sheet("Defect Log"), run, report_cfg or ReportConfig())
    _matrix_sheet(wb.create_sheet("Traceability Matrix"), run)
    _evidence_sheet(wb.create_sheet("Evidence Index"), run)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(output_path)
    log.info("Excel traceability matrix written: %s", output_path)
    return output_path


def _header_row(ws: Worksheet, headers: list[str], widths: list[int]) -> None:
    for col, (header, width) in enumerate(zip(headers, widths, strict=True), 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(vertical="center")
        ws.column_dimensions[get_column_letter(col)].width = width
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def _summary_sheet(ws: Worksheet, run: ValidationRun) -> None:
    ws.title = "Executive Summary"
    summary = run.summary
    manifest = run.manifest
    rows: list[tuple[str, object]] = [
        ("Run ID", manifest.run_id),
        ("System Under Test", manifest.base_url),
        ("Specification", manifest.spec_document),
        ("Specification SHA-256", manifest.spec_sha256),
        ("Tool Version", manifest.tool_version),
        ("Started (UTC)", manifest.started_at),
        ("Finished (UTC)", manifest.finished_at),
        ("Operator", manifest.operator or "automated"),
        ("", ""),
        ("Total Requirements", summary.total),
        ("Passed", summary.passed),
        ("Failed", summary.failed),
        ("Warnings", summary.warnings),
        ("Errors", summary.errors),
        ("Not Tested", summary.not_tested),
        ("Pass Rate (executed)", f"{summary.pass_rate}%"),
    ]
    ws.column_dimensions["A"].width = 28
    ws.column_dimensions["B"].width = 90
    for idx, (label, value) in enumerate(rows, 1):
        ws.cell(row=idx, column=1, value=label).font = Font(bold=True)
        ws.cell(row=idx, column=2, value=str(value) if value != "" else "")
    for status, count_row in ((Status.PASS, 11), (Status.FAIL, 12), (Status.WARNING, 13),
                              (Status.ERROR, 14), (Status.NOT_TESTED, 15)):
        ws.cell(row=count_row, column=2).fill = _STATUS_FILL[status]


def _matrix_sheet(ws: Worksheet, run: ValidationRun) -> None:
    headers = [
        "Requirement ID", "Category", "Requirement Description", "Expected Result",
        "Actual Result", "Status", "Details", "Page URL", "Evidence Location",
        "Spec Source (page / method)", "Timestamp (UTC)", "Duration (ms)",
    ]
    _header_row(ws, headers, [15, 14, 50, 40, 50, 12, 40, 40, 45, 22, 21, 13])
    row = 2
    for req in run.requirements:
        res = run.result_for(req.id)
        status = res.status if res else Status.NOT_TESTED
        values = [
            req.id,
            req.category.value,
            req.requirement,
            req.expected,
            res.actual if res else "Not executed",
            status.value,
            res.details if res else "",
            res.page_url if res else "",
            "\n".join(e.path for e in res.evidence) if res else "",
            f"{req.source.document} p.{req.source.page_number} / {req.source.extraction_method}",
            res.finished_at if res else "",
            res.duration_ms if res else "",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _THIN
            cell.alignment = Alignment(vertical="top", wrap_text=col in (3, 4, 5, 7, 9))
        status_cell = ws.cell(row=row, column=6)
        status_cell.fill = _STATUS_FILL[status]
        status_cell.font = _STATUS_FONT[status]
        row += 1


# Requirement category -> tracker "Defect Type"
_DEFECT_TYPE = {
    "Content": "Content", "Metadata": "Content", "Accessibility": "Accessibility",
    "Visual": "UI/Design", "Responsive": "UI/Design", "Performance": "Performance",
}


def _defect_log_sheet(ws: Worksheet, run: ValidationRun, cfg: ReportConfig) -> None:
    """QA defect-tracker format (one row per Fail/Warning/Error result)."""
    headers = [
        "Defect No", "QA Name", "QA Date", "Phase", "Environment", "Round of Testing",
        "Defect URL", "Device", "Section / Page Name", "Steps / Defect Description",
        "Defect Type", "Defect Category", "Severity", "QA Status", "DEV Status",
        "Reference/Verified File", "Defect Screenshot", "Developer Comments",
        "QC Comments", "DPM Comments",
    ]
    widths = [10, 20, 12, 8, 12, 9, 40, 15, 24, 60, 12, 14, 9, 10, 10, 28, 42, 20, 20, 20]
    _header_row(ws, headers, widths)

    row = 2
    defect_no = 0
    for res in run.results:
        if res.status not in (Status.FAIL, Status.WARNING, Status.ERROR):
            continue
        req = run.requirements.get(res.requirement_id)
        defect_no += 1
        section = (urlparse(res.page_url).path.strip("/") or "Home") if res.page_url else ""
        device = "Desktop Chrome"
        blob = f"{res.actual} {res.details}"
        for hint in ("iPhone 14", "iPad"):
            if hint in blob:
                device = hint
                break
        requirement_text = req.requirement if req else ""
        description = (
            (f"{requirement_text}\n" if requirement_text else "")
            + f"Expected: {res.expected}\nActual: {res.actual}"
            + (f"\nDetails: {res.details}" if res.details else "")
        )
        screenshots = "\n".join(
            e.path for e in res.evidence
            if e.kind in (EvidenceKind.SCREENSHOT, EvidenceKind.ELEMENT_SCREENSHOT, EvidenceKind.VISUAL_DIFF)
        ) or "\n".join(e.path for e in res.evidence)
        category = req.category.value if req else "General"
        values = [
            f"DEF-{defect_no:03d}",
            cfg.qa_name,
            (res.finished_at or run.manifest.finished_at)[:10],
            cfg.phase,
            cfg.environment,
            cfg.round_of_testing,
            res.page_url,
            device,
            section,
            description,
            _DEFECT_TYPE.get(category, "Functional"),
            category,
            "Major" if res.status in (Status.FAIL, Status.ERROR) else "Minor",
            "Open",
            "New",
            f"{req.source.document} (p.{req.source.page_number})" if req else run.manifest.spec_document,
            screenshots,
            "", "", "",
        ]
        for col, value in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = _THIN
            cell.alignment = Alignment(vertical="top", wrap_text=col in (10, 16, 17))
        sev_cell = ws.cell(row=row, column=13)
        sev_cell.fill = _STATUS_FILL[Status.FAIL if values[12] == "Major" else Status.WARNING]
        row += 1


def _evidence_sheet(ws: Worksheet, run: ValidationRun) -> None:
    headers = ["Requirement ID", "Kind", "Path", "Description", "SHA-256", "Captured (UTC)"]
    _header_row(ws, headers, [15, 18, 50, 45, 40, 21])
    row = 2
    for res in run.results:
        for ev in res.evidence:
            values = [res.requirement_id, ev.kind.value, ev.path, ev.description, ev.sha256, ev.captured_at]
            for col, value in enumerate(values, 1):
                cell = ws.cell(row=row, column=col, value=value)
                cell.border = _THIN
                cell.alignment = Alignment(vertical="top")
            row += 1
